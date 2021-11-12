from pathlib import Path
import pandas as pd  # pip install pandas
from openpyxl import load_workbook  # pip install openpyxl

input_dir = Path.cwd() / "Excel_Files"

#--- 1st EXAMPLE
parts = []
for path in list(input_dir.rglob("*.xls*")):
    part = pd.read_excel(path)
    files.append(part)

df = pd.concat(parts)
output_dir = Path.cwd() / "MasterFile"
output_dir.mkdir(exist_ok=True)
df.to_excel(output_dir / "masterfile.xlsx", index=False)


#--- 2nd EXAMPLE
for path in list(input_dir.rglob("*.xls*")):
    wb = load_workbook(filename=path)
    ws = wb["Sheet1"]
    ws["A1"] = "New ID"
    output_dir = Path.cwd() / "New ID"
    output_dir.mkdir(exist_ok=True)
    wb.save(output_dir / path.name)
